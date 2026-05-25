import os
import smtplib
from datetime import datetime
from email.message import EmailMessage

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook


# Load environment variables from .env file
load_dotenv()

# Excel file path
CUSTOMERS_FILE = "data/customers.xlsx"

# Email settings
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")


def create_excel_file_if_not_exists():
    """Create customers Excel file with headers if it does not exist."""

    os.makedirs("data", exist_ok=True)

    if not os.path.exists(CUSTOMERS_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"

        sheet.append([
            "Date",
            "Customer Name",
            "Customer Email",
            "Company",
            "Service Interest"
        ])

        workbook.save(CUSTOMERS_FILE)


def get_customer_info():
    """Collect customer information from terminal input."""

    customer_name = input("Customer name: ").strip()
    customer_email = input("Customer email: ").strip()
    company = input("Company: ").strip()
    service_interest = input("Service interest: ").strip()

    if not customer_name or not customer_email:
        raise ValueError("Customer name and email are required.")

    return {
        "name": customer_name,
        "email": customer_email,
        "company": company,
        "service_interest": service_interest,
    }


def save_customer_to_excel(customer):
    """Save customer information to Excel file."""

    workbook = load_workbook(CUSTOMERS_FILE)
    sheet = workbook["Customers"]

    sheet.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        customer["name"],
        customer["email"],
        customer["company"],
        customer["service_interest"],
    ])

    workbook.save(CUSTOMERS_FILE)


def build_email_message(customer):
    """Build automatic welcome email message."""

    message = EmailMessage()

    message["Subject"] = "Welcome to Our CRM Automation"
    message["From"] = EMAIL_ADDRESS
    message["To"] = customer["email"]

    body = f"""
Hello {customer["name"]},

Thank you for your interest in our services.

We have received your information:
Company: {customer["company"]}
Service Interest: {customer["service_interest"]}

Our team will contact you soon.

Best regards,
CRM Automation Team
"""

    message.set_content(body)

    return message


def send_email(customer):
    """Send automatic email to customer using SMTP."""

    if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
        raise ValueError("Email credentials are missing in .env file.")

    message = build_email_message(customer)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(message)


def main():
    """Main CRM automation workflow."""

    try:
        create_excel_file_if_not_exists()

        customer = get_customer_info()

        save_customer_to_excel(customer)

        send_email(customer)

        print("Customer saved successfully.")
        print("Welcome email sent successfully.")

    except Exception as error:
        print(f"Error: {error}")


if __name__ == "__main__":
    main()
